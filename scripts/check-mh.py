# -*- coding: utf-8 -*-
import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('Copy of Tam Quốc Huyễn Tưởng.xlsx', data_only=True)
ws = wb['Võ tướng Ngụy']

for r, name in [(17, 'Trương Liêu'), (25, 'Trương Cáp')]:
    print(f'\n=== {name} (R{r}) — cột H (Mệnh Hồn Đỏ) ===')
    v = ws[f'H{r}'].value
    print(v if v else '(rỗng)')
