# -*- coding: utf-8 -*-
import sys
sys.path.insert(0, 'scripts')
from importlib import import_module
m = import_module('import-from-excel')
from openpyxl import load_workbook

wb = load_workbook('Copy of Tam Quốc Huyễn Tưởng.xlsx', data_only=True)
ws = wb['Võ tướng Ngụy']

with open('scripts/_debug.txt', 'w', encoding='utf-8') as f:
    for col in ['E', 'F', 'G', 'H']:
        text = ws[f'{col}25'].value
        f.write(f'\n##### Trương Cáp - Cột {col} #####\n')
        cleaned, fx = m.extract_local_effects(text or '')
        f.write(f'Extracted {len(fx)} effects:\n')
        for e in fx:
            f.write(f"  - [{e['name']}]: {e['description'][:80]}\n")
        f.write(f'\n--- Cleaned text ({len(cleaned)} chars) ---\n')
        f.write(cleaned[:800] + '\n')
print('Wrote to scripts/_debug.txt')
