# -*- coding: utf-8 -*-
"""Kiểm tra tiền-import: file JSON tồn tại, ảnh tồn tại, dữ liệu hiện tại."""
import io, json, os, re, sys, unicodedata
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def slugify(name):
    s = unicodedata.normalize('NFD', name)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd').replace('Đ', 'D')
    s = re.sub(r'[^a-zA-Z0-9]+', '_', s).strip('_').lower()
    return s

wb = load_workbook('Copy of Tam Quốc Huyễn Tưởng.xlsx', data_only=True)
ws = wb['Võ tướng Ngụy']

print(f"{'R':>3} {'Tên':<32} {'id':<28} {'JSON':<5} {'IMG':<5} {'skills':<7} {'profession':<10} {'name_cn'}")
print('-' * 110)
for r in range(2, 27):
    name = ws[f'B{r}'].value
    if not name:
        continue
    clean = re.sub(r'\s*\([^)]*\)\s*$', '', name).strip()
    hero_id = slugify(clean)
    json_file = f'data/heroes/{hero_id}.json'
    img_path = f'assets/images/{hero_id.replace("_", "-")}.png'
    json_exists = os.path.exists(json_file)
    img_exists = os.path.exists(img_path)
    existing_skills = 0
    existing_prof = ''
    existing_namecn = ''
    if json_exists:
        h = json.loads(open(json_file, encoding='utf-8').read())
        existing_skills = sum(1 for s in (h.get('skills') or []) if (s.get('description') or '').strip())
        existing_prof = h.get('profession', '')
        existing_namecn = h.get('name_cn', '')
    print(f"{r:>3} {clean:<32} {hero_id:<28} {str(json_exists):<5} {str(img_exists):<5} {existing_skills:<7} {existing_prof:<10} {existing_namecn}")
