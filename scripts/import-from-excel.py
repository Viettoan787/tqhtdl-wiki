# -*- coding: utf-8 -*-
"""
Nhập tướng từ Excel "Copy of Tam Quốc Huyễn Tưởng.xlsx" sang JSON wiki.

Mặc định: DRY RUN (in JSON ra stdout, không ghi file).
Dùng --write để ghi thật vào data/heroes/<id>.json.

Sheet hỗ trợ: "Võ tướng Ngụy" (đầu tiên). Các sheet khác có thể thêm sau.
"""
import argparse
import io
import json
import re
import sys
import unicodedata
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / 'Copy of Tam Quốc Huyễn Tưởng.xlsx'
HEROES_DIR = ROOT / 'data' / 'heroes'

QUALITY_MAP = {
    'UR': 'UR',
    'SSR': 'SSR', 'Thần Tướng': 'SSR',
    'SR': 'SR', 'Danh Tướng': 'SR',
    'R': 'R', 'Lương Tướng': 'R',
}

COUNTRY_BY_SHEET = {
    'Võ tướng Ngụy': ('nguy', 'Ngụy'),
    'Võ tướng Thục': ('thuc', 'Thục'),
    'Võ tướng Ngô': ('ngo', 'Ngô'),
    'Võ tướng Quần': ('quan-hung', 'Quần Hùng'),
}


def slugify(name: str) -> str:
    """'Nhạc Tiến' -> 'nhac_tien'. Bỏ dấu, lowercase, gạch dưới."""
    s = unicodedata.normalize('NFD', name)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd').replace('Đ', 'D')
    s = re.sub(r'[^a-zA-Z0-9]+', '_', s).strip('_').lower()
    return s


def clean_name(raw: str) -> str:
    """'Nhạc Tiến (Yue Jin)' -> 'Nhạc Tiến'."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', (raw or '').strip())


def normalize_text(s: str) -> str:
    if s is None:
        return ''
    s = str(s).replace('\r\n', '\n').replace('\r', '\n').strip()
    return s


# Regex các marker chính của skill — không được nhầm với "[Tên hiệu ứng]"
SECTION_BREAK_RE = re.compile(
    r'(?im)^\s*(?:'
    r'Lv\s*\d|Vô\s*[Ss]ong|DUYÊN|Duyên|MỆNH|Mệnh|PHỔ|Phổ|NỘ|Nộ|KỸ|Kỹ|Bị\s*động|'
    r'Lược\s*hữu|Lư\s*hỏa|Lô\s*hỏa|Thần\s*hồ|Đánh\s*thường|Sóc|'
    r'\(\s*Chú\s*thích\s*hiệu\s*ứng\s*\)?|Chú\s*thích\s*hiệu\s*ứng'
    r')'
)


def extract_local_effects(text: str):
    """Rút các định nghĩa hiệu ứng cục bộ '[Tên]:' hoặc '[Tên]\\n<mô tả>' khỏi text.
    Trả (cleaned_text, [{name, description}])."""
    if not text:
        return text, []
    lines = text.split('\n')
    heading_re = re.compile(r'^\s*\[([^\]\n]{1,40})\]\s*(?::\s*(.*))?\s*$')

    headings = []  # (line_idx, name, inline)
    for i, line in enumerate(lines):
        m = heading_re.match(line)
        if m:
            name = m.group(1).strip()
            inline = (m.group(2) or '').strip()
            headings.append((i, name, inline))

    if not headings:
        return text, []

    used = set()
    effects = []
    for idx, (h_idx, name, inline) in enumerate(headings):
        next_h_idx = headings[idx + 1][0] if idx + 1 < len(headings) else len(lines)
        body_lines = [inline] if inline else []
        candidate_used = {h_idx}
        for j in range(h_idx + 1, next_h_idx):
            if SECTION_BREAK_RE.match(lines[j]):
                break
            body_lines.append(lines[j])
            candidate_used.add(j)
        body = '\n'.join(body_lines).strip()
        # Chỉ coi là định nghĩa nếu có mô tả thực sự (không phải 1 dòng đơn lẻ trong câu)
        if not body or len(body) < 5:
            continue
        # Tránh nhầm với token inline: heading phải đứng đầu một section
        # (đứng sau dòng trống hoặc đầu file, không nằm cùng dòng câu)
        if h_idx > 0 and lines[h_idx - 1].strip() and not SECTION_BREAK_RE.match(lines[h_idx - 1]):
            # Nếu dòng trước cũng có nội dung (không phải break), kiểm tra xem có phải đoạn liên tục từ sentence không
            # Heuristic: nếu dòng trước kết thúc bằng `.` hoặc `;` hoặc `:`, coi như đã hết câu
            prev = lines[h_idx - 1].rstrip()
            if not prev.endswith(('.', ';', ':', '!', '?')):
                continue
        used |= candidate_used
        effects.append({'name': name, 'description': body})

    cleaned_lines = [line for i, line in enumerate(lines) if i not in used]
    cleaned = '\n'.join(cleaned_lines).strip()
    return cleaned, effects


# ---------- Parsers ----------

def parse_active_skills(text: str):
    """Cột E: trả về (pho_cong, no_cong, bi_dong_3, bi_dong_5).
    Hỗ trợ nhiều format: PHỔ CÔNG, Phổ công, Đánh thường, NỘ CÔNG, Kỹ năng Nộ, KỸ NĂNG BỊ ĐỘNG..."""
    text = normalize_text(text)
    out = {'pho_cong': None, 'no_cong': None, 'bi_dong_3_sao': None, 'bi_dong_5_sao': None}

    # Marker: bắt nhiều biến thể, group(0) = nhóm marker, group('rest') = phần còn lại trên cùng dòng (có thể là tên)
    section_re = re.compile(
        r'(?im)^[\s\W]*'
        r'(?P<kind>'
        r'PHỔ\s*CÔNG|Phổ\s*công|Đánh\s*thường|'
        r'NỘ\s*CÔNG|Nộ\s*công|Kỹ\s*năng\s*Nộ|'
        r'KỸ\s*NĂNG\s*BỊ\s*ĐỘNG|Kỹ\s*năng\s*Bị\s*Động|Kỹ\s*năng\s*bị\s*động'
        r')'
        r'[ \t]*(?P<rest>[^\n]*)'
    )
    markers = list(section_re.finditer(text))
    if not markers:
        return out

    sections = []
    for i, m in enumerate(markers):
        end = markers[i + 1].start() if i + 1 < len(markers) else len(text)
        sections.append({
            'kind_raw': m.group('kind'),
            'inline': m.group('rest').strip(),
            'body': text[m.end():end].strip(),
        })

    for sec in sections:
        kind = re.sub(r'\s+', ' ', sec['kind_raw'].strip().lower())
        if kind in ('phổ công', 'đánh thường'):
            out['pho_cong'] = extract_named_skill(sec['inline'], sec['body'])
        elif kind in ('nộ công', 'kỹ năng nộ'):
            out['no_cong'] = extract_named_skill(sec['inline'], sec['body'])
        elif 'bị động' in kind:
            for skill in parse_passives_v2(sec['body']):
                key = f"bi_dong_{skill['star']}_sao"
                if key in out:
                    out[key] = {'name': skill['name'], 'description': skill['description']}
    return out


def extract_named_skill(inline: str, body: str):
    """Phổ Công / Nộ Công có thể có tên cùng dòng marker (vd 'PHỔ CÔNG Thiết Quyền')
    hoặc tên ở dòng đầu tiên của body dạng '<Tên>: <mô tả>' (vd 'Ngôn Linh: Gây sát thương...')."""
    if inline:
        return {'name': inline, 'description': body}
    # Thử tách "<Tên>: <mô tả>" từ dòng đầu của body
    first_line_end = body.find('\n')
    first_line = body[:first_line_end] if first_line_end != -1 else body
    rest = body[first_line_end + 1:] if first_line_end != -1 else ''
    if ':' in first_line:
        name_part, desc_part = first_line.split(':', 1)
        # Tránh trường hợp dòng đầu là một câu mô tả dài có ":" giữa câu
        if len(name_part.strip()) <= 40 and '\n' not in name_part:
            full_desc = (desc_part.strip() + '\n' + rest).strip() if rest else desc_part.strip()
            return {'name': name_part.strip(), 'description': full_desc}
    # Không tìm thấy tên — để trống
    return {'name': '', 'description': body}


def parse_passives_v2(body: str):
    """Bị động: hỗ trợ 2 format chính:
    - '<Tên> (Võ tướng N sao kích hoạt): <mô tả>'
    - 'N Sao - <Tên>: <mô tả>' / '<N> sao - <Tên>: <mô tả>'
    """
    skills = []
    # Format 1: "(Võ tướng N sao kích hoạt)"
    p1 = re.compile(r'\(Võ\s*tướng\s*(\d)\s*sao\s*kích\s*hoạt\)\s*:?\s*', re.IGNORECASE)
    matches = list(p1.finditer(body))
    if matches:
        for i, m in enumerate(matches):
            prev_end = matches[i - 1].end() if i > 0 else 0
            before = body[prev_end:m.start()].rstrip()
            name = before.rsplit('\n', 1)[-1].strip()
            if i + 1 < len(matches):
                after = body[m.end():matches[i + 1].start()]
                last_nl = after.rfind('\n')
                desc = after[:last_nl].strip() if last_nl != -1 else after.strip()
            else:
                desc = body[m.end():].strip()
            skills.append({'name': name, 'star': int(m.group(1)), 'description': desc})
        return skills

    # Format 2: "N Sao - <Tên>: <mô tả>" (Tuân Úc style)
    p2 = re.compile(r'(?im)^[\s\W]*?(\d)\s*[Ss]ao\s*[-–—]\s*([^:\n]+?)\s*:\s*', re.MULTILINE)
    matches = list(p2.finditer(body))
    for i, m in enumerate(matches):
        end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
        skills.append({
            'name': m.group(2).strip(),
            'star': int(m.group(1)),
            'description': body[m.end():end].strip(),
        })
    return skills


VS_PATTERN = re.compile(r'(?im)^Vô\s*[Ss]ong\s*(?:Bậc\s*)?(1|3|5)\b[^\n:]*:?\s*')

def parse_vs_duyen(text: str):
    """Cột F: trả (vo_song[1/3/5], duyen_phan)."""
    text = normalize_text(text)
    out = {'vo_song': {}, 'duyen_phan': None}

    # Tách Vô Song
    matches = list(VS_PATTERN.finditer(text))
    for i, m in enumerate(matches):
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[m.end():end].strip()
        out['vo_song'][int(m.group(1))] = body

    # Duyên phận: hỗ trợ "DUYÊN PHẬN <tên>", "Duyên phận <tên>", "Duyên: <tên>"
    dp = re.search(r'(?im)^Duyên(?:\s+Phận)?\s*[:—–\-]?\s*([^\n]+)\n(.+)$', text, re.DOTALL)
    if dp:
        out['duyen_phan'] = {'name': dp.group(1).strip(), 'description': dp.group(2).strip()}

    # Trim Vô Song body để không kéo theo Duyên Phận
    if out['duyen_phan']:
        for k in list(out['vo_song'].keys()):
            v = out['vo_song'][k]
            cut = re.search(r'(?im)^Duyên(?:\s+Phận)?\b', v)
            if cut:
                out['vo_song'][k] = v[:cut.start()].strip()
    return out


def parse_huyen_vu(text: str):
    """Cột G: trả về (name, stages dict, local_effects list)."""
    text = normalize_text(text)
    if not text:
        return None
    name = ''
    name_match = re.search(r'^\s*\[([^\]\n]+)\]', text)
    if name_match:
        name = vietnamese_titlecase(name_match.group(1).strip())

    stages = {'Lược Hữu Tiểu Thành': '', 'Lư Hỏa Thuần Thanh': '', 'Thần Hồ Kỳ Kỹ': ''}
    # Hỗ trợ 2 format:
    # - "Lược hữu tiểu thành: ..." / "Lư hỏa thuần thanh: ..." / "Thần hồ kỳ kỹ: ..."
    # - "Lv1: ..." (= Lược) / "Lv2: ..." (= Lư) / "Lv3: ..." (= Thần)
    stage_pattern = re.compile(
        r'(?im)^(Lược\s*hữu\s*tiểu\s*thành|Lư\s*hỏa\s*thuần\s*thanh|Lô\s*hỏa\s*thuần\s*thanh|Thần\s*hồ\s*kỳ\s*kỹ|Lv\s*1|Lv\s*2|Lv\s*3)\s*[:\-–—]?\s*'
    )
    matches = list(stage_pattern.finditer(text))
    # Phần sau stage cuối có thể chứa "(Chú thích hiệu ứng)" - cần cắt
    end_marker = re.search(r'(?im)^\(?\s*Chú thích hiệu ứng\s*\)?', text)

    lv_to_stage = {'lv1': 'Lược Hữu Tiểu Thành', 'lv2': 'Lư Hỏa Thuần Thanh', 'lv3': 'Thần Hồ Kỳ Kỹ'}
    for i, m in enumerate(matches):
        if i + 1 < len(matches):
            end = matches[i + 1].start()
        elif end_marker and end_marker.start() > m.end():
            end = end_marker.start()
        else:
            end = len(text)
        body = text[m.end():end].strip()
        key_raw = re.sub(r'\s+', '', m.group(1).strip().lower())
        if key_raw in lv_to_stage:
            stages[lv_to_stage[key_raw]] = body
        elif 'lượchữu' in key_raw or 'lược' in key_raw:
            stages['Lược Hữu Tiểu Thành'] = body
        elif 'hỏa' in key_raw:
            stages['Lư Hỏa Thuần Thanh'] = body
        elif 'thần' in key_raw:
            stages['Thần Hồ Kỳ Kỹ'] = body

    return {'name': name, 'stages': stages}


def parse_menh_hon(text: str):
    """Cột H: trả về (base_name, descs[1/3/5])."""
    text = normalize_text(text)
    if not text:
        return None
    base_name = ''
    name_match = re.search(r'^\s*\[([^\]\n]+)\]', text)
    if name_match:
        base_name = vietnamese_titlecase(name_match.group(1).strip())

    # Marker bậc: "[Mệnh hồn giác tỉnh nhất giai]" hoặc "Mệnh Hồn Giác Tỉnh Bậc 1" hoặc "Lv1:"
    rank_map = [
        (re.compile(r'(?im)\[?\s*Mệnh\s*hồn\s*giác\s*tỉnh\s*nhất\s*giai\s*\]?\s*:?\s*'), 1),
        (re.compile(r'(?im)\[?\s*Mệnh\s*hồn\s*giác\s*tỉnh\s*tam\s*giai\s*\]?\s*:?\s*'), 3),
        (re.compile(r'(?im)\[?\s*Mệnh\s*hồn\s*giác\s*tỉnh\s*ngũ\s*giai\s*\]?\s*:?\s*'), 5),
        (re.compile(r'(?im)Mệnh\s*Hồn\s*Giác\s*Tỉnh\s*Bậc\s*1\s*:?\s*'), 1),
        (re.compile(r'(?im)Mệnh\s*Hồn\s*Giác\s*Tỉnh\s*Bậc\s*3\s*:?\s*'), 3),
        (re.compile(r'(?im)Mệnh\s*Hồn\s*Giác\s*Tỉnh\s*Bậc\s*5\s*:?\s*'), 5),
        (re.compile(r'(?im)Thức\s*tỉnh\s*mệnh\s*hồn\s*bậc\s*1\s*:?\s*'), 1),
        (re.compile(r'(?im)Thức\s*tỉnh\s*mệnh\s*hồn\s*bậc\s*3\s*:?\s*'), 3),
        (re.compile(r'(?im)Thức\s*tỉnh\s*mệnh\s*hồn\s*bậc\s*5\s*:?\s*'), 5),
        (re.compile(r'(?im)^\s*Lv\s*1\s*[:\-–—]?\s*'), 1),
        (re.compile(r'(?im)^\s*Lv\s*2\s*[:\-–—]?\s*'), 3),
        (re.compile(r'(?im)^\s*Lv\s*3\s*[:\-–—]?\s*'), 5),
        (re.compile(r'(?im)^\s*Lv\s*5\s*[:\-–—]?\s*'), 5),
    ]

    # Tìm tất cả marker, gom theo vị trí
    found = []
    for pat, rank in rank_map:
        for m in pat.finditer(text):
            found.append((m.start(), m.end(), rank))
    found.sort()
    # Loại trùng vị trí (ưu tiên match đầu)
    seen_ranks = set()
    cleaned = []
    for s, e, r in found:
        if r in seen_ranks:
            continue
        seen_ranks.add(r)
        cleaned.append((s, e, r))
    cleaned.sort()

    descs = {}
    for i, (s, e, r) in enumerate(cleaned):
        end = cleaned[i + 1][0] if i + 1 < len(cleaned) else len(text)
        descs[r] = text[e:end].strip()

    return {'name': base_name, 'descs': descs}


def vietnamese_titlecase(s: str) -> str:
    """Chuyển ALL CAPS tiếng Việt về Title Case từng từ."""
    parts = s.split()
    return ' '.join(p[:1].upper() + p[1:].lower() for p in parts)


# Đổi tên hiệu ứng theo tướng (sửa thuật ngữ Việt hóa)
PER_HERO_TOKEN_RENAMES = {
    'truong_cap': [('Cạm Bẫy', 'Hãm Tỉnh')],
}


def apply_token_renames(hero: dict):
    pairs = PER_HERO_TOKEN_RENAMES.get(hero.get('id'))
    if not pairs:
        return hero
    for skill in hero.get('skills', []):
        d = skill.get('description', '')
        for old, new in pairs:
            d = d.replace(f'[{old}]', f'[{new}]')
        skill['description'] = d
    return hero


def rename_effect_name(hero_id: str, name: str) -> str:
    pairs = PER_HERO_TOKEN_RENAMES.get(hero_id) or []
    for old, new in pairs:
        if name == old:
            return new
    return name


# ---------- Build hero JSON ----------

def label_with_name(prefix: str, name: str) -> str:
    name = (name or '').strip()
    return f'{prefix} — {name}' if name else prefix


def build_hero(row: dict, country: str, faction: str, existing: dict | None):
    name_raw = row['name']
    clean = clean_name(name_raw)
    hero_id = existing['id'] if existing else slugify(clean)
    quality_raw = (row['quality'] or '').strip()
    quality = QUALITY_MAP.get(quality_raw, existing['quality'] if existing else 'SSR')

    # Tách hiệu ứng local từ mọi ô trước khi parse
    cell_effects = []
    for col_key in ('col_E', 'col_F', 'col_G', 'col_H'):
        text = normalize_text(row.get(col_key) or '')
        cleaned, fx = extract_local_effects(text)
        row[col_key] = cleaned
        cell_effects.extend(fx)

    hero = {
        'id': hero_id,
        'name': clean,
        'name_cn': existing.get('name_cn', '') if existing else '',
        'type': 'normal',
        'country': country,
        'faction': faction,
        'role': 'Võ Tướng',
        'profession': existing.get('profession', 'Chưa rõ') if existing else 'Chưa rõ',
        'quality': quality,
        'image': existing.get('image', f'/assets/images/{hero_id.replace("_", "-")}.png') if existing else f'/assets/images/{hero_id.replace("_", "-")}.png',
        'description': existing.get('description', f'Võ tướng phe {faction}.') if existing else f'Võ tướng phe {faction}.',
    }
    if existing and existing.get('avatar'):
        hero['avatar'] = existing['avatar']

    skills = []
    actives = parse_active_skills(row['col_E'])
    if actives['pho_cong'] and actives['pho_cong']['description']:
        skills.append({
            'id': f'{hero_id}_pho_cong',
            'name': label_with_name('Phổ Công', actives['pho_cong']['name']),
            'type': 'active', 'cooldown': None,
            'description': actives['pho_cong']['description'],
        })
    if actives['no_cong'] and actives['no_cong']['description']:
        skills.append({
            'id': f'{hero_id}_no_cong',
            'name': label_with_name('Nộ Công', actives['no_cong']['name']),
            'type': 'active', 'cooldown': None,
            'description': actives['no_cong']['description'],
        })
    for star_key, label in (('bi_dong_3_sao', 'Bị động 3 sao'), ('bi_dong_5_sao', 'Bị động 5 sao')):
        s = actives[star_key]
        if s and s['description']:
            skills.append({
                'id': f'{hero_id}_{star_key}',
                'name': label_with_name(label, s['name']),
                'type': 'passive', 'cooldown': None,
                'description': s['description'],
            })

    vsdp = parse_vs_duyen(row['col_F'])
    for lv in (1, 3, 5):
        body = vsdp['vo_song'].get(lv)
        if body:
            skills.append({
                'id': f'{hero_id}_vo_song_{lv}',
                'name': f'Vô Song {lv}',
                'type': 'ultimate', 'cooldown': None,
                'description': body,
            })
    if vsdp['duyen_phan'] and vsdp['duyen_phan']['description']:
        skills.append({
            'id': f'{hero_id}_duyen_phan_1',
            'name': label_with_name('Duyên Phận', vsdp['duyen_phan']['name']),
            'type': 'passive', 'cooldown': None,
            'description': vsdp['duyen_phan']['description'],
        })

    new_local_effects = cell_effects[:]
    hv = parse_huyen_vu(row['col_G'])
    if hv:
        parts = []
        for stage_name in ('Lược Hữu Tiểu Thành', 'Lư Hỏa Thuần Thanh', 'Thần Hồ Kỳ Kỹ'):
            if hv['stages'].get(stage_name):
                parts.append(f"**{stage_name}**: {hv['stages'][stage_name]}")
        if parts:
            skills.append({
                'id': f'{hero_id}_huyen_vu',
                'name': label_with_name('Huyễn Vũ', hv['name']),
                'type': 'passive', 'cooldown': None,
                'description': '\n\n'.join(parts),
            })

    mh = parse_menh_hon(row['col_H'])
    if mh and mh['descs']:
        for lv in (1, 3, 5):
            body = mh['descs'].get(lv)
            if body:
                if mh['name']:
                    name = f"Mệnh Hồn Đỏ — {mh['name']} {lv}"
                else:
                    name = f"Mệnh Hồn Đỏ {lv}"
                skills.append({
                    'id': f'{hero_id}_menh_hon_do_{lv}',
                    'name': name.replace('  ', ' ').strip(),
                    'type': 'ultimate', 'cooldown': None,
                    'description': body,
                })

    hero['skills'] = skills
    apply_token_renames(hero)
    return hero, new_local_effects


def load_existing_hero(hero_id: str):
    f = HEROES_DIR / f'{hero_id}.json'
    if f.exists():
        return json.loads(f.read_text(encoding='utf-8'))
    # tên file dạng dấu nối khác (vd thai-van-co.json)
    alt = HEROES_DIR / f'{hero_id.replace("_", "-")}.json'
    if alt.exists():
        return json.loads(alt.read_text(encoding='utf-8'))
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--sheet', default='Võ tướng Ngụy')
    parser.add_argument('--row', type=int, action='append', help='Một dòng (1-indexed). Lặp lại để chạy nhiều dòng. Bỏ qua: tất cả.')
    parser.add_argument('--write', action='store_true', help='Ghi thật vào data/heroes/. Mặc định là dry-run.')
    parser.add_argument('--force', action='store_true', help='Ghi đè cả tướng đã có dữ liệu thật.')
    parser.add_argument('--update-effects', action='store_true', help='Merge các hiệu ứng local vào data/effects.json (chỉ thêm, không ghi đè).')
    args = parser.parse_args()

    if args.sheet not in COUNTRY_BY_SHEET:
        print(f'Sheet không hỗ trợ: {args.sheet}', file=sys.stderr)
        sys.exit(2)
    country, faction = COUNTRY_BY_SHEET[args.sheet]

    wb = load_workbook(EXCEL, data_only=True)
    ws = wb[args.sheet]

    rows_to_process = args.row if args.row else list(range(2, ws.max_row + 1))
    summary = []
    aggregated_effects = {}

    for r in rows_to_process:
        name_cell = ws[f'B{r}'].value
        if not name_cell:
            continue
        row = {
            'name': name_cell,
            'quality': ws[f'D{r}'].value,
            'col_E': ws[f'E{r}'].value,
            'col_F': ws[f'F{r}'].value,
            'col_G': ws[f'G{r}'].value,
            'col_H': ws[f'H{r}'].value,
        }
        clean = clean_name(row['name'])
        candidate_id = slugify(clean)
        existing = load_existing_hero(candidate_id)
        # Nếu file đã có data thật (>0 skill có nội dung), đánh dấu để cẩn thận
        existing_has_real = bool(existing and any(s.get('description') for s in (existing.get('skills') or [])))

        hero, local_effects = build_hero(row, country, faction, existing)
        for fx in local_effects:
            final_name = rename_effect_name(hero['id'], fx['name'])
            aggregated_effects.setdefault(final_name, fx['description'])

        skill_count = len(hero['skills'])
        action = 'WRITE'
        if not args.write:
            action = 'DRY-RUN'
        elif existing_has_real and not args.force:
            action = 'SKIP (đã có data thật, dùng --force để ghi đè)'
        else:
            target = HEROES_DIR / f'{hero["id"]}.json'
            target.write_text(json.dumps(hero, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

        summary.append({
            'row': r, 'name': clean, 'id': hero['id'],
            'quality': hero['quality'], 'skills': skill_count,
            'existing_has_real': existing_has_real, 'action': action,
        })

        if args.row:
            print(json.dumps(hero, ensure_ascii=False, indent=2))

    print('\n=== SUMMARY ===')
    for s in summary:
        print(json.dumps(s, ensure_ascii=False))

    if aggregated_effects:
        print('\n=== HIỆU ỨNG LOCAL TÌM THẤY ===')
        for name, desc in aggregated_effects.items():
            print(f"- [{name}]: {desc[:120]}{'...' if len(desc) > 120 else ''}")

        if args.update_effects and args.write:
            effects_path = ROOT / 'data' / 'effects.json'
            existing_effects = json.loads(effects_path.read_text(encoding='utf-8'))
            existing_names = {e.get('name') for e in existing_effects}
            existing_ids = {e.get('id') for e in existing_effects}
            added = 0
            for name, desc in aggregated_effects.items():
                if name in existing_names:
                    continue
                effect_id = slugify(name)
                base_id = effect_id
                n = 2
                while effect_id in existing_ids:
                    effect_id = f'{base_id}_{n}'
                    n += 1
                existing_effects.append({'id': effect_id, 'name': name, 'description': desc})
                existing_ids.add(effect_id)
                existing_names.add(name)
                added += 1
            if added:
                effects_path.write_text(json.dumps(existing_effects, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
                print(f'\n=> Đã thêm {added} hiệu ứng mới vào data/effects.json.')
            else:
                print('\n=> Tất cả hiệu ứng đã có trong data/effects.json (bỏ qua).')


if __name__ == '__main__':
    main()
