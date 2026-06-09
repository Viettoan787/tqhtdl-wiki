# -*- coding: utf-8 -*-
"""In nội dung đầy đủ vài tướng để hiểu format ô."""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook('Copy of Tam Quốc Huyễn Tưởng.xlsx', data_only=True)
ws = wb['Võ tướng Ngụy']

COLS = {
  'A': 'STT', 'B': 'Tên', 'D': 'Phẩm', 'E': 'Kỹ năng 5 sao',
  'F': 'Vô Song / Duyên', 'G': 'Huyễn Vũ', 'H': 'Mệnh Hồn Đỏ', 'I': 'Cơ chế lõi'
}

# In Nhạc Tiến (R4) và Tào Phi (R10) — Tào Phi đã có dữ liệu thật để đối chiếu.
for r in [4, 10]:
    print(f"\n{'='*100}\nROW {r}")
    for col, label in COLS.items():
        v = ws[f'{col}{r}'].value
        v = '(rỗng)' if v is None else str(v)
        print(f"\n--- [{col}] {label} ---\n{v}")
